import openpyxl as xl
import algo
import random,time,sys,os

merged_cells={}
c=[] # 小组列表
v=[]

def shuffle(file1, file2, method, output):
    """主函数"""
    global merged_cells,c,v
    with open(file1,encoding='utf8') as f:
        # 随时更改此处
        text=f.read()
        students=[]
        genders=[]
        for i in text.split('\n'):
            j=i.split(' ')
            students.append(j[0])
            genders.append(j[1])
    book=xl.load_workbook(file2)
    sheet=book.active

    # 识别小组
    dim=sheet.calculate_dimension()
    dim_cell=sheet[dim.split(':')[-1]]
    dimx=dim_cell.row
    dimy=dim_cell.column
    ls1=list(sheet.merged_cells)
    for i in ls1:
        y1,x1,y2,x2=i.bounds
        cell1=sheet.cell(x1,y1)
        for j in range(x1,x2+1):
            for k in range(y1,y2+1):
                merged_cells[sheet.cell(j,k)]=cell1
    v=[[False for j in range(dimy+1)] for i in range(dimx+1)]
    c=[]
    for i in range(1,dimx+1): # 为什么这个b索引是从1开始的？？？
        for j in range(1,dimy+1):
            if not v[i][j] and if_bordered(sheet.cell(i,j)):
                c.append([])
                dfs(i,j,dimx,dimy,sheet)
    # print(c)

    # 数据处理
    n=len(c)
    m=[]
    pre_boys=[]
    pre_girls=[]
    total_boys=genders.count('男')
    total_girls=genders.count('女')
    pre_names=[]
    for i in c:
        m.append(len(i))
        pre_boys.append(0)
        pre_girls.append(0)
        for j in i:
            val=j.value
            if val in students:
                g=genders[students.index(val)]
                if g=='男':
                    pre_boys[-1]+=1
                if g=='女':
                    pre_girls[-1]+=1
                pre_names.append(val)
            elif val:
                m[-1]-=1
    # print(n,m,pre_boys,pre_girls,sep='\n')

    # 调用算法
    print(type(method))
    if method==0:
        res=algo.allocate_groups(n,m,pre_boys,pre_girls,total_boys,total_girls)
    if method==1:
        res=algo.allocate_groups_separated(n,m,pre_boys,pre_girls,total_boys,total_girls)
    if method==2:
        res=algo.allocate_groups_random(n,m,pre_boys,pre_girls,total_boys,total_girls)
    # print(res)

    # 分配座位
    for i in range(len(c)):
        c[i].sort(key=lambda x:x.row*dimy+x.column) # 按单元格索引排序
    seed=time.time()
    random.seed(seed)
    random.shuffle(students)
    random.seed(seed)
    random.shuffle(genders)
    boys=filter(lambda x:genders[students.index(x)]=='男',students)
    girls=filter(lambda x:genders[students.index(x)]=='女',students)
    for i in range(n):
        # print(res[i])
        # print('i',i)
        boys_cnt,girls_cnt=res[i][0]-pre_boys[i],res[i][1]-pre_girls[i]
        gender_ls=[0]*boys_cnt+[1]*girls_cnt
        random.shuffle(gender_ls)
        p=0 # 座位
        q=0 # 学生
        while q<(boys_cnt+girls_cnt):
            # print('q',q)
            if gender_ls[q]:
                target=next(girls)
            else:
                target=next(boys)
            if target in pre_names:
                continue
            # print(target)
            while c[i][p].value:
                p+=1
                # print('p',p)
            c[i][p].value=target
            q+=1

    book.save(output)

def dfs(x,y,dimx,dimy,sheet):
    global c,v
    cell=sheet.cell(x,y)
    v[x][y]=True
    if type(cell)==xl.cell.cell.MergedCell:
        if cell==merged_cells[cell]:
            c[-1].append(cell)
        cell=merged_cells[cell]
    else:
        c[-1].append(cell)
    for i in [[0,1],[1,0],[0,-1],[-1,0]]:
        x1,y1=x+i[0],y+i[1]
        if x1<1 or x1>dimx or y1<1 or y1>dimy:
            continue
        if v[x1][y1]:
            continue
        cell1=sheet.cell(x1,y1)
        if type(cell1)==xl.cell.cell.MergedCell:
            cell1=merged_cells[cell1]
        if not if_bordered(cell1) or cell.fill.fgColor!=cell1.fill.fgColor:
            continue
        dfs(x1,y1,dimx,dimy,sheet)

def if_bordered(cell):
    return cell.border.left.style and cell.border.right.style and cell.border.top.style and cell.border.bottom.style

def res_path(relative_path):
    """获取打包后资源文件的绝对路径"""
    if hasattr(sys, '_MEIPASS'):
        # 如果是打包后的环境
        base_path = sys._MEIPASS
    else:
        # 开发环境，直接使用当前路径
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

if __name__=='__main__':
    shuffle('test.txt','test.xlsx',2,'result.xlsx')
